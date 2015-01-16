COLORS = ['b', 'g', 'r', 'c', 'm', 'b']

def plot(points, links=True, dots=True):
    """If dots is set to True, draws dots on the points. If links is True,
    draws links between points. If both are false nothing is drawn.
    
    Points should be in the form of ((xlist, ylist), (xlist2, ylist2), ...)."""
    
    import matplotlib.pyplot as plt
    
    if len(points) > COLORS:
        print "Cannot draw more than %d lines, you specified %d" % (len(COLORS), len(point))
        assert False
    
    # Plot links
    if links:
        for line_idx, pts in enumerate(points):
            # Use None-separated list as explained on http://exnumerus.blogspot.nl/2011/02/how-to-quickly-plot-multiple-line.html
            xlist, ylist = pts
            
            plt.plot(xlist, ylist, color=COLORS[line_idx])
        
    # Plot points
    if dots:
        for line_idx, pts in enumerate(points):
            xlist, ylist = pts
            plt.scatter(xlist, ylist, color=COLORS[line_idx])

    plt.show()

def find_in_worksheet(ws, row, col, multiple=False):
    """Finds a value in a worksheet such that the column it is in has prefix col
    and the row that it is in starts with prefix row. Col and row may be tuples
    to indicate that its first few cells should have those value. Specify None
    in these tuples to specify that the value of that specific value does not
    matter. Specify an integer for col or row if the column or row is known.
    
    Returns the bottom-right-most cell that adheres to the given prefixes. Returns None
    if no cell was found that adheres to the given prefixes.
    
    If multiple is True, returns a (possibly empty) list."""

    if multiple:
        row_nr = []
        col_nr = []
    else:
        row_nr = None
        col_nr = None
    
    if type(row) == str:
        row = (row,)
    if type(col) == str:
        col = (col,)
    
    # Find the row
    if type(row) == int:
        if multiple:
            row_nr = [row]
        else:
            row_nr = row
    else:
        for row_idx, row_cells in enumerate(ws.rows):
            for idx, cell in enumerate(row_cells):
                if idx >= len(row):
                    if multiple:
                        row_nr.append(row_idx)
                    else:
                        row_nr = row_idx
                    break
                elif row[idx] is None:
                    continue
                elif row[idx] != cell.value:
                    break
        
    # Find the column
    if type(col) == int:
        if multiple:
            col_nr = [col]
        else:
            col_nr = col
    else:
        for col_idx, col_cells in enumerate(ws.columns):
            for idx, cell in enumerate(col_cells):
                if idx >= len(col):
                    if multiple:
                        col_nr.append(col_idx)
                    else:
                        col_nr = col_idx
                    break
                elif col[idx] is None:
                    continue
                elif col[idx] != cell.value:
                    break

    if multiple:
        result = []
        for r in row_nr:
            for c in col_nr:
                result.append(ws.cell(None, r, c))
        return result
    else:
        if col_nr and row_nr:
            return ws.cell(None, row_nr, col_nr)
        else:
            return None

def get_metric(ws, dataset, algorithm, dilation_ratio, metric):
    """ .. """
    
    v = lambda v: list(c.value for c in v)

    if dataset is None:
        set_sizes = find_in_worksheet(ws, (None, None, algorithm, dilation_ratio), "#Vertices", True)
        metric_cells = find_in_worksheet(ws, (None, None, algorithm, dilation_ratio), metric, True)
        return v(set_sizes), v(metric_cells)
        
    elif algorithm is None:
        algorithms = find_in_worksheet(ws, (dataset, None, None, dilation_ratio), "Algorithm", True)
        metric_cells = find_in_worksheet(ws, (dataset, None, None, dilation_ratio), metric, True)
        return v(algorithms), v(metric_cells)
        
    elif dilation_ratio is None:
        ratio_cells = find_in_worksheet(ws, (dataset, None, algorithm), "Required dilation ratio", True)
        metric_cells = find_in_worksheet(ws, (dataset, None, algorithm), metric, True)
        return v(ratio_cells), v(metric_cells)
        
    else:
        print "One of dataset, algorithm or dilation_ratio should be None"
        assert False
        
if __name__ == "__main__":
    import openpyxl
    
    wb = openpyxl.load_workbook("experiment log.xlsx")
    ws = wb.get_active_sheet()
