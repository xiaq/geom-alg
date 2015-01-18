import time
import openpyxl
import sys

from greedy_spanner import greedy_spanner, greedy_theta_spanner
from yao import yao_graph
from theta import theta_graph
from datatypes import Graph
from WSPD import wspd_spanner

from data.train_stations import train_stations_by_label_luxembourg, train_stations_by_label_netherlands

DO_ONLY_GRAPHS = []
DO_ONLY_ALGOS = []
DO_ONLY_DILATIONS = []

# First generate the graphs
GRAPHS = {}
#Random graphs
for n in range(25, 351, 25):
    GRAPHS["Random %d" % n] = Graph.from_data_challenge("data/random %d.txt" % n, removeDuplicates = True)[0]
#Train stations NL
GRAPHS["Train stations NL"] = Graph.datapoints_graph([v for k, v in train_stations_by_label_netherlands.iteritems() if v is not None], removeDuplicates = True)
#Train stations Luxembourg
GRAPHS["Train stations LUX"] = Graph.datapoints_graph([v for k, v in train_stations_by_label_luxembourg.iteritems() if v is not None], removeDuplicates = True)
#Our own data challenge
GRAPHS["Our data challenge"] = Graph.from_data_challenge("data/data challenge.txt", removeDuplicates = True)[0]
for n in range(1,7): #(1..6)
    GRAPHS["Data challenge %d" % n] = Graph.from_data_challenge("data/B%d.txt" % n, removeDuplicates = True)[0]

if DO_ONLY_GRAPHS:
    GRAPHS = {k: v for k, v in GRAPHS.iteritems() if k in DO_ONLY_GRAPHS}

# Algorithms
ALGOS = {"Greedy": greedy_spanner, "Greedy Theta": greedy_theta_spanner, "Yao": yao_graph, "Theta": theta_graph, "WSPD": wspd_spanner}
ALGO_MAX_SIZE = {"Greedy": 450, "Greedy Theta": 3000, "Yao": 3000, "Theta": 3000, "WSPD": 2000}
if DO_ONLY_ALGOS:
    ALGOS = {k: v for k, v in ALGOS.iteritems() if k in DO_ONLY_ALGOS}

# Dilation ratios
RATIOS = (1.1, 1.2, 1.3, 1.5, 2, 3, 5)
if DO_ONLY_DILATIONS:
    RATIOS = [d for d in RATIOS if d in DO_ONLY_DILATIONS]
    
LOG_CONTAINER = "experiment log.xlsx"
LOG_NAME = "Run of %s" % time.strftime("%a %m-%d %H.%M.%S")

# See if we're running from IDLE, if not, open a stderr text file
if "idlelib" not in sys.modules:
    sys.stderr = open("stderr.txt", "a")

# Prepare Excel file
log_wb = openpyxl.load_workbook(LOG_CONTAINER)
log_ws = log_wb.create_sheet(0, LOG_NAME)
for col, txt in enumerate(("Graph", "#Vertices", "Algorithm", "Required dilation ratio", "Actual dilation ratio", "#Edges", "Total edge weight", "Max edge degree", "Diameter", "#Intersections", "Running time")):
    log_ws.cell(None, 0, col).value = txt

run_nr = 1  
n_graphs_done = 0

for g_name, g in GRAPHS.iteritems():
    for algo_name, algo in ALGOS.iteritems():
        if g.n_vertices() > ALGO_MAX_SIZE[algo_name]:
            print "Skipping algorithm %s for graph %s" % (algo_name, g_name)
            continue
            
        for ratio in RATIOS:
            t_start = time.clock()
            algo(g, ratio)
            t_elapsed = time.clock() - t_start
            
            for col, txt in enumerate((g_name, g.n_vertices(), algo_name, ratio, g.dilation_ratio(), g.n_edges(), g.weight(), g.max_edge_degree(), 0, 0, t_elapsed)): # Replace 0, 0 with g.diameter() and the intersections function to also compute these metrics
                log_ws.cell(None, run_nr, col).value = txt
            
            g.clear_edges()
            run_nr += 1
            print "Completed", g_name, algo_name, ratio
            
        log_wb.save(LOG_CONTAINER)
    
    n_graphs_done += 1
    print "COMPLETED PROCESSING %d OUT OF %d GRAPHS" % (n_graphs_done, len(GRAPHS))
